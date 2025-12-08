"""
Estrattore Centrale Rischi - Web App
Carica un PDF della CR e ottieni un Excel con le tabelle separate.
"""
import streamlit as st
import pdfplumber
import re
import io
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

st.set_page_config(
    page_title="Estrattore Centrale Rischi",
    page_icon="📊",
    layout="wide"
)

MESI_IT = {
    'gennaio': '01', 'febbraio': '02', 'marzo': '03', 'aprile': '04',
    'maggio': '05', 'giugno': '06', 'luglio': '07', 'agosto': '08',
    'settembre': '09', 'ottobre': '10', 'novembre': '11', 'dicembre': '12'
}

SEZIONI_MAP = {
    'Crediti Per Cassa': 'CREDITI_PER_CASSA',
    'Crediti Di Firma': 'CREDITI_DI_FIRMA',
    'Garanzie Ricevute': 'GARANZIE_RICEVUTE',
    'Derivati Finanziari': 'DERIVATI_FINANZIARI',
    'Sezione Informativa': 'SEZIONE_INFORMATIVA',
    'Informazioni Sui Garanti': 'INFO_GARANTI',
    'Informazioni Sui Debitori Ceduti': 'INFO_DEBITORI_CEDUTI',
    'Richieste Di Informazione': 'RICHIESTE_INFO',
    'Informazioni Generali': 'INFO_GENERALI',
}

CATEGORIE_NORMALIZE = {
    'SOFFERENZESOFFERENZE': 'SOFFERENZE',
    'RISCHI AUTOLIQUIDANTI': 'RISCHI_AUTOLIQUIDANTI',
    'RISCHI A SCADENZA': 'RISCHI_A_SCADENZA',
    'RISCHI A REVOCA': 'RISCHI_A_REVOCA',
}

CAMPI_NUMERICI = [
    'accordato', 'accordato_operativo', 'utilizzato', 'saldo_medio',
    'importo', 'importo_garantito', 'valore_garanzia', 'valore_intrinseco',
    'valore_nominale_del_credito_ceduto', 'ammontare_della_garanzia_rilasciata',
    'garantito', 'ruolo_affidato'
]

VALID_HEADERS = {
    "categoria", "localizzazione", "durata", "durata_originaria", "durata_residua",
    "divisa", "import_export", "tipo_attivita", "tipo_garanzia", "ruolo_affidato",
    "stato_rapporto", "accordato", "accordato_operativo", "utilizzato", 
    "saldo_medio", "importo", "importo_garantito", "garante", "valore_garanzia",
    "tipo_evento", "data_evento", "note", "valore_intrinseco",
    "valore_nominale_del_credito_ceduto", "nominativo_richiesto",
    "tipo_richiesta_di_informazione", "data_della_richiesta_di_informazione",
    "causale_della_richiesta", "periodo_richiesto", "periodo_validita",
    "intermediario_che_ha_effettuato_la_richiesta", "ceduto",
    "variabili_di_classificazione", "classi_di_dato",
    "ammontare_della_garanzia_rilasciata", "garantito",
    "descrizione_causale", "evento_cancellato",
}

BLACKLIST_HEADERS = {"a", "da", "ria", "catego", "categorie", "classi_dato"}
HEADER_CORRECTIONS = {"periodo_validit": "periodo_validita", "tipo_attivit": "tipo_attivita"}
MAX_HEADER_LENGTH = 60

RE_DATA_RIF = re.compile(r"DATA DI RIFERIMENTO:\s*([a-zA-Z]+ \d{4})", re.IGNORECASE)
RE_INTERMEDIARIO = re.compile(r"Intermediario:\s*(.+?)(?:\n|$)", re.IGNORECASE)


# ============================================================================
# FUNZIONI DI NORMALIZZAZIONE
# ============================================================================

def normalizza_data_riferimento(data_str):
    if not data_str:
        return None
    data_str = data_str.strip().lower()
    match = re.match(r'([a-z]+)\s+(\d{4})', data_str)
    if match:
        mese_str, anno = match.groups()
        mese_num = MESI_IT.get(mese_str)
        if mese_num:
            return f"{anno}-{mese_num}"
    return data_str


def normalizza_data_gg_mm_yyyy(data_str):
    if not data_str:
        return None
    match = re.match(r'(\d{2})/(\d{2})/(\d{4})', data_str.strip())
    if match:
        gg, mm, yyyy = match.groups()
        return f"{yyyy}-{mm}-{gg}"
    return data_str


def normalizza_importo(val_str):
    if not val_str:
        return None
    val_str = str(val_str).strip().replace(' ', '')
    if not val_str or val_str == '-':
        return None
    if val_str.isdigit():
        return float(val_str)
    if ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    else:
        parts = val_str.split('.')
        if len(parts) > 1 and all(len(p) == 3 for p in parts[1:]):
            val_str = val_str.replace('.', '')
    try:
        return float(val_str)
    except ValueError:
        return None


def normalizza_categoria(cat_str):
    if not cat_str:
        return ''
    cat_str = cat_str.strip().upper()
    if cat_str in CATEGORIE_NORMALIZE:
        return CATEGORIE_NORMALIZE[cat_str]
    return re.sub(r'[^A-Z0-9_]', '', re.sub(r'\s+', '_', cat_str))


def normalizza_sezione(sez_str):
    if not sez_str:
        return ''
    return SEZIONI_MAP.get(sez_str, sez_str.upper().replace(' ', '_'))


def split_periodo_richiesto(periodo_str):
    if not periodo_str:
        return None, None
    match = re.match(r'([A-Za-z]+)\s+(\d{4})\s*-\s*([A-Za-z]+)\s+(\d{4})', periodo_str, re.IGNORECASE)
    if match:
        mese_da, anno_da, mese_a, anno_a = match.groups()
        da = f"{anno_da}-{MESI_IT.get(mese_da.lower(), '00')}"
        a = f"{anno_a}-{MESI_IT.get(mese_a.lower(), '00')}"
        return da, a
    return None, None


# ============================================================================
# FUNZIONI DI ESTRAZIONE PDF
# ============================================================================

def normalizza_header(h):
    if h is None:
        return None
    h = re.sub(r'[\n\r]+', ' ', h)
    h = re.sub(r'\s+', ' ', h).strip()
    if not h or len(h) > MAX_HEADER_LENGTH:
        return None
    h = h.lower().replace(" ", "_")
    h = re.sub(r"[^a-z0-9_]", "", h)
    if h in BLACKLIST_HEADERS or len(h) <= 1:
        return None
    return HEADER_CORRECTIONS.get(h, h)


def is_valid_table(table):
    if not table or len(table) < 2:
        return False
    header = table[0]
    if not header:
        return False
    valid_cols = [c for c in header if c and str(c).strip()]
    if len(valid_cols) < 2:
        return False
    first_header = str(header[0]).lower().strip() if header[0] else ""
    if first_header.startswith(("ntermediario", "intermediario:")):
        return False
    norm_headers = [normalizza_header(h) for h in header]
    known_count = sum(1 for h in norm_headers if h and h in VALID_HEADERS)
    valid_count = sum(1 for h in norm_headers if h)
    return known_count >= 1 or valid_count >= 3


def trova_sezione_da_header(header_row):
    if not header_row:
        return None
    norm_headers = set(normalizza_header(h) for h in header_row if h)
    
    if 'garante' in norm_headers and 'valore_garanzia' in norm_headers:
        return "Informazioni Sui Garanti"
    if norm_headers == {'categoria', 'localizzazione', 'stato_rapporto', 'importo'}:
        return "Sezione Informativa"
    if 'ammontare_della_garanzia_rilasciata' in norm_headers:
        return "Crediti Di Firma"
    if 'valore_intrinseco' in norm_headers:
        return "Derivati Finanziari"
    if 'garantito' in norm_headers and 'tipo_garanzia' in norm_headers and 'garante' not in norm_headers:
        return "Garanzie Ricevute"
    if 'tipo_richiesta_di_informazione' in norm_headers or \
       'data_della_richiesta_di_informazione' in norm_headers or \
       'nominativo_richiesto' in norm_headers:
        return "Richieste Di Informazione"
    if 'ceduto' in norm_headers and 'valore_nominale_del_credito_ceduto' in norm_headers:
        return "Informazioni Sui Debitori Ceduti"
    if 'tipo_evento' in norm_headers and 'data_evento' in norm_headers:
        return "Sezione Informativa"
    if 'variabili_di_classificazione' in norm_headers:
        return "Informazioni Generali"
    if 'accordato' in norm_headers or 'utilizzato' in norm_headers:
        return "Crediti Per Cassa"
    return None


def pulisci_valore(value):
    if value is None:
        return ""
    value = str(value)
    value = re.sub(r'[\n\r]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def estrai_da_pdf(pdf_file, progress_callback=None):
    rows = []
    dynamic_fields = set()
    
    with pdfplumber.open(pdf_file) as pdf:
        contesto = {"data_riferimento": None, "intermediario": None}
        total_pages = len(pdf.pages)
        
        for page_idx, page in enumerate(pdf.pages):
            if progress_callback:
                progress_callback((page_idx + 1) / total_pages)
            
            page_number = page_idx + 1
            text = page.extract_text() or ""
            
            m_data = RE_DATA_RIF.search(text)
            if m_data:
                contesto["data_riferimento"] = m_data.group(1).strip()
            m_int = RE_INTERMEDIARIO.search(text)
            if m_int:
                contesto["intermediario"] = m_int.group(1).strip()
            
            tables = page.extract_tables()
            if not tables:
                continue
            
            for table in tables:
                if not is_valid_table(table):
                    continue
                
                header_row = table[0]
                sezione = trova_sezione_da_header(header_row)
                norm_headers = [normalizza_header(h) for h in header_row]
                
                for h in norm_headers:
                    if h:
                        dynamic_fields.add(h)
                
                for r in table[1:]:
                    if all((c is None or str(c).strip() == "") for c in r):
                        continue
                    
                    row_dict = {
                        "pagina": page_number,
                        "data_riferimento": contesto["data_riferimento"],
                        "intermediario": contesto["intermediario"],
                        "sezione": sezione,
                    }
                    
                    for idx, cell in enumerate(r):
                        col_name = norm_headers[idx] if idx < len(norm_headers) else None
                        if col_name:
                            row_dict[col_name] = pulisci_valore(cell)
                    
                    rows.append(row_dict)
    
    return rows, dynamic_fields


def normalizza_riga(row):
    result = row.copy()
    
    if result.get('data_riferimento'):
        result['data_riferimento'] = normalizza_data_riferimento(result['data_riferimento'])
    
    if result.get('sezione'):
        result['sezione_cod'] = normalizza_sezione(result['sezione'])
    
    if result.get('categoria'):
        result['categoria_cod'] = normalizza_categoria(result['categoria'])
    
    for campo in CAMPI_NUMERICI:
        if result.get(campo):
            result[f'{campo}_num'] = normalizza_importo(result[campo])
    
    for campo in ['data_della_richiesta_di_informazione', 'data_evento']:
        if result.get(campo):
            result[campo] = normalizza_data_gg_mm_yyyy(result[campo])
    
    if result.get('periodo_richiesto'):
        da, a = split_periodo_richiesto(result['periodo_richiesto'])
        result['periodo_richiesto_da'] = da
        result['periodo_richiesto_a'] = a
    
    return result


def rimuovi_duplicati(rows):
    seen = set()
    unique = []
    key_fields = ['data_riferimento', 'intermediario', 'sezione', 'categoria',
                  'accordato', 'utilizzato', 'importo', 'garante', 'valore_garanzia']
    for row in rows:
        key = tuple(row.get(f, '') for f in key_fields)
        if key not in seen:
            seen.add(key)
            unique.append(row)
    return unique


def split_tabelle(rows):
    tabelle = {
        'crediti_cassa': [], 'crediti_firma': [], 'derivati': [],
        'garanti': [], 'garanzie_ricevute': [], 'sezione_informativa': [],
        'debitori_ceduti': [], 'richieste_info': [], 'info_generali': [],
    }
    sezione_map = {
        'CREDITI_PER_CASSA': 'crediti_cassa', 'CREDITI_DI_FIRMA': 'crediti_firma',
        'DERIVATI_FINANZIARI': 'derivati', 'INFO_GARANTI': 'garanti',
        'GARANZIE_RICEVUTE': 'garanzie_ricevute', 'SEZIONE_INFORMATIVA': 'sezione_informativa',
        'INFO_DEBITORI_CEDUTI': 'debitori_ceduti', 'RICHIESTE_INFO': 'richieste_info',
        'INFO_GENERALI': 'info_generali',
    }
    for row in rows:
        sez_cod = row.get('sezione_cod', '')
        target = sezione_map.get(sez_cod, 'crediti_cassa')
        tabelle[target].append(row)
    return tabelle


# ============================================================================
# GENERAZIONE EXCEL
# ============================================================================

def genera_excel(tabelle, intestatario=None):
    fogli_config = [
        ("crediti_cassa", "Crediti Cassa"),
        ("garanti", "Garanti"),
        ("sezione_informativa", "Sezione Informativa"),
        ("derivati", "Derivati"),
        ("richieste_info", "Richieste Info"),
        ("garanzie_ricevute", "Garanzie Ricevute"),
        ("debitori_ceduti", "Debitori Ceduti"),
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fill_light = PatternFill("solid", fgColor="D6DCE4")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Foglio riepilogo
    ws_summary = wb.create_sheet(title="Riepilogo", index=0)
    ws_summary['A1'] = f"CENTRALE RISCHI - {intestatario or 'Estratto'}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    ws_summary['A3'] = "Data estrazione:"
    ws_summary['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    ws_summary['A6'] = "Tabella"
    ws_summary['B6'] = "Righe"
    ws_summary['A6'].font = header_font
    ws_summary['A6'].fill = header_fill
    ws_summary['B6'].font = header_font
    ws_summary['B6'].fill = header_fill
    
    row_idx = 7
    totale = 0
    
    for key, nome in fogli_config:
        righe = tabelle.get(key, [])
        if not righe:
            continue
        
        ws_summary.cell(row=row_idx, column=1, value=nome)
        ws_summary.cell(row=row_idx, column=2, value=len(righe))
        totale += len(righe)
        row_idx += 1
        
        # Crea foglio dati
        ws = wb.create_sheet(title=nome)
        
        if righe:
            headers = list(righe[0].keys())
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for r_idx, row in enumerate(righe, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    value = row.get(header, '')
                    
                    if header.endswith('_num') and value:
                        try:
                            cell.value = float(value)
                            cell.number_format = '#,##0'
                        except:
                            cell.value = value
                    else:
                        cell.value = value
                    
                    cell.border = border
                    if r_idx % 2 == 0:
                        cell.fill = num_fill_light
            
            # Auto-width
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row in righe[:100]:
                    val = str(row.get(header, ''))
                    max_len = max(max_len, len(val))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
            
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
    
    ws_summary.cell(row=row_idx, column=1, value="TOTALE").font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2, value=totale).font = Font(bold=True)
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================================
# UI STREAMLIT
# ============================================================================

st.title("📊 Estrattore Centrale Rischi")
st.markdown("""
Carica un PDF della Centrale Rischi di Banca d'Italia e ottieni un file Excel 
con i dati strutturati in tabelle separate.
""")

uploaded_file = st.file_uploader(
    "Carica il PDF della Centrale Rischi",
    type=['pdf'],
    help="Seleziona il file PDF scaricato dalla Centrale Rischi"
)

if uploaded_file is not None:
    st.info(f"📄 File caricato: **{uploaded_file.name}** ({uploaded_file.size / 1024:.1f} KB)")
    
    if st.button("🚀 Estrai dati", type="primary"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            # Estrazione
            status_text.text("📖 Lettura PDF...")
            rows, _ = estrai_da_pdf(uploaded_file, progress_callback=progress_bar.progress)
            
            status_text.text("🔄 Normalizzazione dati...")
            rows = [normalizza_riga(r) for r in rows]
            
            original_count = len(rows)
            rows = rimuovi_duplicati(rows)
            duplicates = original_count - len(rows)
            
            status_text.text("📊 Organizzazione tabelle...")
            tabelle = split_tabelle(rows)
            
            # Estrai intestatario dal nome file
            intestatario = uploaded_file.name.replace('.pdf', '').replace('_', ' ')
            
            status_text.text("📝 Generazione Excel...")
            excel_buffer = genera_excel(tabelle, intestatario)
            
            progress_bar.progress(100)
            status_text.text("✅ Completato!")
            
            # Statistiche
            st.success(f"Estratte **{len(rows)}** righe ({duplicates} duplicati rimossi)")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### 📈 Riepilogo tabelle")
                for nome, righe in tabelle.items():
                    if righe and nome != 'info_generali':
                        st.write(f"- **{nome.replace('_', ' ').title()}**: {len(righe)} righe")
            
            with col2:
                st.markdown("### ⬇️ Download")
                output_filename = uploaded_file.name.replace('.pdf', '_estratto.xlsx')
                st.download_button(
                    label="📥 Scarica Excel",
                    data=excel_buffer,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
        
        except Exception as e:
            st.error(f"❌ Errore durante l'elaborazione: {str(e)}")
            st.exception(e)

st.markdown("---")
st.markdown("""
<div style="text-align: center; color: gray; font-size: 0.8em;">
    Estrattore Centrale Rischi v5 | Sviluppato con ❤️ da Kitzanos
</div>
""", unsafe_allow_html=True)
